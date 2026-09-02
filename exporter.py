import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime

def export_attendance_to_excel(session, records, file_path):
    """
    Generates a professional Excel (.xlsx) report for an attendance session.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Summary"
    
    # Title Banner
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"AI ATTENDANCE REPORT: {session.session_title.upper()}"
    title_cell.font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45

    # Session Meta
    ws.cell(row=2, column=1, value=f"Class: {session.class_name}").font = Font(bold=True)
    ws.cell(row=2, column=4, value=f"Professor: {session.created_by_teacher_name}").font = Font(bold=True)
    ws.cell(row=2, column=6, value=f"Date: {session.created_at.strftime('%Y-%m-%d %H:%M')}").font = Font(bold=True)
    
    # Headers
    headers = ["S.No", "Roll No", "Student Name", "Class", "Status", "Marking Method", "Professor Override / Marked By"]
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='38BDF8')
    thin_border = Border(
        left=Side(style='thin', color='334155'),
        right=Side(style='thin', color='334155'),
        top=Side(style='thin', color='334155'),
        bottom=Side(style='thin', color='334155')
    )

    ws.append([]) # Empty row 3
    ws.append(headers) # Row 4
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    # Data Rows
    present_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    present_font = Font(color='15803D', bold=True)
    absent_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    absent_font = Font(color='B91C1C', bold=True)

    for idx, r in enumerate(records, 1):
        row_num = 4 + idx
        student_name = r.student.name if r.student else "N/A"
        roll_no = r.student.roll_no if r.student else "N/A"
        class_name = r.student.class_name if r.student else "N/A"
        marked_by = r.marked_by_teacher_name if r.marked_by_teacher_name else "System AI"

        row_data = [
            idx,
            roll_no,
            student_name,
            class_name,
            r.status,
            r.marking_method.replace('_', ' '),
            marked_by
        ]
        ws.append(row_data)
        
        # Style row
        status_cell = ws.cell(row=row_num, column=5)
        if r.status == 'PRESENT':
            status_cell.fill = present_fill
            status_cell.font = present_font
        else:
            status_cell.fill = absent_fill
            status_cell.font = absent_font
            
        for c in range(1, 8):
            cell = ws.cell(row=row_num, column=c)
            cell.border = thin_border
            if c in [1, 2, 4, 5]:
                cell.alignment = Alignment(horizontal='center')

    # Auto-fit Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(file_path)
    return file_path

def export_attendance_to_pdf(session, records, file_path):
    """
    Generates a publication-quality PDF report for an attendance session using ReportLab.
    """
    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter,
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#0F172A'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        textColor=colors.HexColor('#475569'),
        alignment=1,
        spaceAfter=20
    )

    cell_bold = ParagraphStyle('CellBold', fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#0F172A'))
    cell_norm = ParagraphStyle('CellNorm', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#334155'))
    status_pres = ParagraphStyle('StatPres', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#16A34A'))
    status_abs = ParagraphStyle('StatAbs', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#DC2626'))

    elements = []
    
    # Title & Header
    elements.append(Paragraph(f"AI Attendance Report: {session.session_title}", title_style))
    elements.append(Paragraph(f"Class: <b>{session.class_name}</b> | Professor: <b>{session.created_by_teacher_name}</b> | Date: <b>{session.created_at.strftime('%Y-%m-%d %H:%M')}</b>", subtitle_style))
    
    # Summary Statistics Box
    total_st = len(records)
    present_cnt = len([r for r in records if r.status == 'PRESENT'])
    absent_cnt = len([r for r in records if r.status == 'ABSENT'])
    percentage = round((present_cnt / total_st * 100), 1) if total_st > 0 else 0.0

    stat_data = [
        [
            Paragraph(f"<b>Total Students:</b> {total_st}", cell_norm),
            Paragraph(f"<b>Present:</b> <font color='#16A34A'>{present_cnt}</font>", cell_norm),
            Paragraph(f"<b>Absent:</b> <font color='#DC2626'>{absent_cnt}</font>", cell_norm),
            Paragraph(f"<b>Attendance Rate:</b> <b>{percentage}%</b>", cell_norm)
        ]
    ]
    t_stat = Table(stat_data, colWidths=[130, 130, 130, 150])
    t_stat.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F1F5F9')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(t_stat)
    elements.append(Spacer(1, 20))

    # Main Records Table
    table_data = [
        [
            Paragraph("S.No", cell_bold),
            Paragraph("Roll No", cell_bold),
            Paragraph("Student Name", cell_bold),
            Paragraph("Status", cell_bold),
            Paragraph("Method", cell_bold),
            Paragraph("Teacher Override / Marked By", cell_bold)
        ]
    ]

    for idx, r in enumerate(records, 1):
        s_name = r.student.name if r.student else "N/A"
        r_no = r.student.roll_no if r.student else "N/A"
        st_style = status_pres if r.status == 'PRESENT' else status_abs
        m_by = r.marked_by_teacher_name if r.marked_by_teacher_name else "System AI"

        table_data.append([
            Paragraph(str(idx), cell_norm),
            Paragraph(r_no, cell_norm),
            Paragraph(s_name, cell_norm),
            Paragraph(r.status, st_style),
            Paragraph(r.marking_method.replace('_', ' '), cell_norm),
            Paragraph(m_by, cell_norm)
        ])

    t_records = Table(table_data, colWidths=[40, 70, 160, 70, 90, 110])
    t_records.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    
    # Header row text color white fix
    for i in range(6):
        t_records._cellvalues[0][i] = Paragraph(
            t_records._cellvalues[0][i].text,
            ParagraphStyle('HeaderStyle', parent=cell_bold, textColor=colors.white)
        )

    elements.append(t_records)
    doc.build(elements)
    return file_path
